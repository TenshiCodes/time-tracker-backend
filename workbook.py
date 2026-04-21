from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import timedelta
from zoneinfo import ZoneInfo
from collections import defaultdict, OrderedDict


def build_timesheet_wb(projects, time_entries, tz="UTC"):
    daily_totals = defaultdict(float)

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    # -----------------------------------
    # ✅ HEADERS FIRST
    # -----------------------------------
    headers = ["Date", "Time {hh:mm}", "Project Number", "Customer", "Labor Type", "Description"]
    ws.append(headers)
    # -----------------------------------
    # ✅ LOOKUP SHEET (for Customer dropdown)
    # -----------------------------------
    ws_lookup = wb.create_sheet(title="Lookup")

    for i, p in enumerate(projects, start=1):
        ws_lookup.cell(row=i, column=1, value=p["name"])  # Customer
        ws_lookup.cell(row=i, column=2, value=p["code"])  # Project Number

    ws_lookup.sheet_state = "hidden"
    max_row_lookup = len(projects)
    # -----------------------------------
    # ✅ PROJECT MAP (code → name)
    # -----------------------------------
    project_map = {p["code"]: p["name"] for p in projects}
    # -----------------------------------
    # ✅ STEP 1: SORT ENTRIES (preserve order)
    # -----------------------------------
    time_entries.sort(key=lambda x: x["clock_in"] or "")
    # -----------------------------------
    # ✅ WRITE DATA
    # -----------------------------------
    grouped = OrderedDict(
        sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1]))
    )

    for row in time_entries:
        start = row["clock_in"]
        end = row["clock_out"]
        job_code = row["job_code"]

        if not start or not end:
            continue  # skip incomplete entries

        start_dt = start.astimezone(ZoneInfo(tz)).replace(tzinfo=None)
        end_dt = end.astimezone(ZoneInfo(tz)).replace(tzinfo=None)

        # Fix bad timestamps
        if end_dt < start_dt:
            diff_hours = (start_dt - end_dt).total_seconds() / 3600
            if diff_hours <= 12:
                end_dt += timedelta(hours=12)
            else:
                end_dt += timedelta(days=1)

        hours = round((end_dt - start_dt).total_seconds() / 3600, 2)

        entry_date = start_dt.date()
        key = (entry_date, job_code)

        if key not in grouped:
            grouped[key] = 0

        grouped[key] += hours
    # -----------------------------------
    # ✅ CHECK DAILY LIMIT (8 HOURS)
    # -----------------------------------
    row_index = 2

    for (entry_date, job_code), hours in grouped.items():
        customer_name = project_map.get(job_code, "")
        remaining_hours = hours

        while remaining_hours > 0.0001:
            current_total = daily_totals[entry_date]

            # -----------------------------------
            # ✅ DETERMINE TIER
            # -----------------------------------
            if current_total < 8:
                tier_limit = 8
                labor_type = "Labor"

            elif current_total < 12:
                tier_limit = 12
                labor_type = "Labor 1.5 X (Bonus OT)"

            else:
                tier_limit = float("inf")
                labor_type = "Labor 2 X (Bonus OT)"

            # -----------------------------------
            # ✅ CALCULATE AVAILABLE SPACE IN TIER
            # -----------------------------------
            available = tier_limit - current_total

            # If already beyond 12, available is infinite
            if tier_limit == float("inf"):
                chunk = remaining_hours
            else:
                chunk = min(remaining_hours, available)

            # -----------------------------------
            # ✅ WRITE ROW
            # -----------------------------------
            ws.cell(row=row_index, column=1, value=entry_date)
            ws.cell(row=row_index, column=1).number_format = "m/d/yyyy"

            ws.cell(row=row_index, column=2, value=chunk / 24)
            ws.cell(row=row_index, column=2).number_format = "[h]:mm"

            ws.cell(row=row_index, column=3, value=job_code)
            ws.cell(row=row_index, column=4, value=customer_name)
            ws.cell(row=row_index, column=5, value=labor_type)

            row_index += 1

            # -----------------------------------
            # ✅ UPDATE TRACKERS
            # -----------------------------------
            daily_totals[entry_date] += chunk
            remaining_hours -= chunk

    # -----------------------------------
    # ✅ CONDITIONAL FORMATTING (EXACT ORIGINAL)
    # -----------------------------------
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    rows = 200

    rules = {
        "B": '=AND($A2<>"", OR(WEEKDAY($A2,2)<6, AND(WEEKDAY($A2,2)>=6, $B2<>"")), B2="")',
        "C": '=AND($A2<>"", OR(WEEKDAY($A2,2)<6, AND(WEEKDAY($A2,2)>=6, $B2<>"")), C2="")',
        "D": '=AND($A2<>"", OR(WEEKDAY($A2,2)<6, AND(WEEKDAY($A2,2)>=6, $B2<>"")), D2="")',
        "E": '=AND($A2<>"", OR(WEEKDAY($A2,2)<6, AND(WEEKDAY($A2,2)>=6, $B2<>"")), E2="")',
        "F": '=AND($A2<>"", OR(WEEKDAY($A2,2)<6, AND(WEEKDAY($A2,2)>=6, $B2<>"")), F2="")',
    }

    for col, formula in rules.items():
        ws.conditional_formatting.add(
            f"{col}2:{col}{rows}",
            FormulaRule(formula=[formula], fill=red_fill)
        )

    blue_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws.conditional_formatting.add(
        "A2:A200",
        FormulaRule(
            formula=['=AND(A2<>"", WEEKDAY(A2,2)>=6)'],
            fill=blue_fill
        )
    )

    # -----------------------------------
    # ✅ HEADER STYLING
    # -----------------------------------
    header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Column widths
    column_widths = [12, 14, 20, 25, 22, 40]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws.row_dimensions[1].height = 25

    # Freeze
    ws.freeze_panes = "A2"

    # -----------------------------------
    # ✅ LABOR DROPDOWN
    # -----------------------------------
    labor_options = [
        "Labor",
        "Labor 1 X (Flex Time)",
        "Labor 1 X (Bonus OT)",
        "Labor 1.5 X (Flex Time)",
        "Labor 1.5 X (Bonus OT)",
        "Labor 2 X (Flex Time)",
        "Labor 2 X (Bonus OT)",
        "Personal Vehicle Ground Travel",
        "Air Travel",
        "PTO (Gusto)",
        "PTO (Embedded Contract)",
        "PTO (Sabbatical)",
        "PTO (Flex Time)",
        "Disability",
        "Company Holiday"
    ]

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(labor_options)}"',
        allow_blank=True
    )
    # -----------------------------------
    # ✅ CUSTOMER DROPDOWN
    # -----------------------------------
    dv_customer = DataValidation(
        type="list",
        formula1=f'=Lookup!$A$1:$A${max_row_lookup}',
        allow_blank=True
    )

    ws.add_data_validation(dv_customer)
    dv_customer.add("D2:D200")

    ws.add_data_validation(dv)
    dv.add("E2:E200")

    return wb